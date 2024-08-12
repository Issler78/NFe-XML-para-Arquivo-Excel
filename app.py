import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import xmltodict
from validate_docbr import CNPJ

# criar arquivo excel
wb = openpyxl.Workbook()
wb.create_sheet('Notas Fiscais')
del wb['Sheet']
nf_page = wb['Notas Fiscais']

# adicionar headers
headers = [
    'Número da Nota',
    'Emissor',
    'CNPJ Emissor',
    'Endereço Emissor',
    'Cidade/UF Emissor',
    'CEP Emissor',
    'Destinatário',
    'CNPJ Destinatário',
    'Endereço Destinatário',
    'Cidade/UF Destinatário',
    'CEP Destinatário',
    'COD Produto',
    'Produto',
    'Valor Unitário',
    'Quantidade Vendida',
    'Valor Total',
    'Peso Líquido'
]
nf_page.append(headers)

def formatar_cnpj(cnpj):
    """Function to return formatted CNPJ

    Args:
        cnpj (string): CNPJ (only numbers)

    Returns:
        string: Formatted CNPJ
    """
    
    cnpj_formatado = CNPJ()
    return cnpj_formatado.mask(cnpj)

# recuperar todos os arquivos xml da pasta NFe
arquivos_xml = os.listdir("Nfe")

for arquivo in arquivos_xml:
    with open(f'NFe/{arquivo}', 'rb') as file:
        # criando um dicionario com os dados do arquivo xml
        nfe_dict = xmltodict.parse(file)
        
        # criar um dict apenas com as infos que precisamos salvar na tabela excel
        nfe_table_dict = {
            'Número da Nota': int(nfe_dict['NFe']['infNFe']['ide']['nNF']),
            'Emissor': nfe_dict['NFe']['infNFe']['emit']['xNome'],
            'CNPJ Emissor': formatar_cnpj(nfe_dict['NFe']['infNFe']['emit']['CNPJ']),
            'Endereço Emissor': f'{nfe_dict['NFe']['infNFe']['emit']['enderEmit']['xLgr']} - {nfe_dict['NFe']['infNFe']['emit']['enderEmit']['nro']}',
            'Cidade/UF Emissor': f'{nfe_dict['NFe']['infNFe']['emit']['enderEmit']['xMun']}/{nfe_dict['NFe']['infNFe']['emit']['enderEmit']['UF']}',
            'CEP Emissor': f'{nfe_dict['NFe']['infNFe']['emit']['enderEmit']['CEP']}',
            'Destinatário': nfe_dict['NFe']['infNFe']['dest']['xNome'],
            'CNPJ Destinatário': formatar_cnpj(nfe_dict['NFe']['infNFe']['dest']['CNPJ']),
            'Endereço Destinatário': f'{nfe_dict['NFe']['infNFe']['dest']['enderDest']['xLgr']} - {nfe_dict['NFe']['infNFe']['dest']['enderDest']['nro']}',
            'Cidade/UF Destinatário': f'{nfe_dict['NFe']['infNFe']['dest']['enderDest']['xMun']}/{nfe_dict['NFe']['infNFe']['dest']['enderDest']['UF']}',
            'CEP Destinatário': f'{nfe_dict['NFe']['infNFe']['dest']['enderDest']['CEP']}',
            'COD Produto': f'{nfe_dict['NFe']['infNFe']['det']['prod']['cProd']}',
            'Produto': f'{nfe_dict['NFe']['infNFe']['det']['prod']['xProd']}',
            'Valor Unitário': f"R${str(nfe_dict['NFe']['infNFe']['det']['prod']['vUnCom']).replace('.', ',')}",
            'Quantidade Vendida': format(float(nfe_dict['NFe']['infNFe']['det']['prod']['qCom']), ".0f"),
            'Valor Total': f'R${str(nfe_dict['NFe']['infNFe']['det']['prod']['vProd']).replace('.', ',')}',
            'Peso Líquido': None
        }
        
        if nfe_dict['NFe']['infNFe']['det']['prod'].get('pesoL'):
            nfe_table_dict['Peso Líquido'] = float(nfe_dict['NFe']['infNFe']['det']['prod']['pesoL'])
    # adicionar na tabela
    values = list(nfe_table_dict.values())
    nf_page.append(values)

# Estilizar o arquivo excel (formatar como tabela)
table_ref = f"A1:Q{nf_page.max_row}"
table = Table(displayName="NotasFiscais", ref=table_ref)

style = TableStyleInfo(
    name="TableStyleMedium2", 
    showFirstColumn=True,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

table.tableStyleInfo = style
nf_page.add_table(table)

# Definindo a largura das colunas
column_widths = {
    'A': 15,  # Número da Nota
    'B': 25,  # Emissor
    'C': 20,  # CNPJ Emissor
    'D': 30,  # Endereço Emissor
    'E': 15,  # Cidade/UF Emissor
    'F': 10,  # CEP Emissor
    'G': 25,  # Destinatário
    'H': 20,  # CNPJ Destinatário
    'I': 30,  # Endereço Destinatário
    'J': 15,  # Cidade/UF Destinatário
    'K': 10,  # CEP Destinatário
    'L': 15,  # COD Produto
    'M': 25,  # Produto
    'N': 15,  # Valor Unitário
    'O': 18,  # Quantidade Vendida
    'P': 15,  # Valor Total
    'Q': 15   # Peso Líquido
}

for col_letter, width in column_widths.items():
    nf_page.column_dimensions[col_letter].width = width

# salvar tabela
wb.save('Excel/Notas_Fiscais.xlsx')
